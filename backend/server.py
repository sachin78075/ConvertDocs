from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Form
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import shutil
import tempfile
import asyncio

# Document conversion imports
from PyPDF2 import PdfReader, PdfWriter, PdfMerger
from pdf2docx import Converter
from docx import Document
from PIL import Image
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
import pytesseract
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from pptx import Presentation
from bs4 import BeautifulSoup

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI(title="ConvertDocs API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Temporary upload directory
UPLOAD_DIR = Path("/tmp/convertdocs_uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Models
class BlogPost(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    slug: str
    excerpt: str
    content: str
    category: str
    author: str = "ConvertDocs Team"
    image_url: Optional[str] = None
    published_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BlogPostCreate(BaseModel):
    title: str
    slug: str
    excerpt: str
    content: str
    category: str
    image_url: Optional[str] = None

class ContactMessage(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    subject: str
    message: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ContactMessageCreate(BaseModel):
    name: str
    email: str
    subject: str
    message: str

# File cleanup task
async def cleanup_old_files():
    """Clean up files older than 1 hour"""
    while True:
        try:
            current_time = datetime.now().timestamp()
            for file_path in UPLOAD_DIR.glob("*"):
                if file_path.is_file():
                    file_age = current_time - file_path.stat().st_mtime
                    if file_age > 3600:  # 1 hour
                        file_path.unlink()
        except Exception as e:
            logging.error(f"Cleanup error: {e}")
        await asyncio.sleep(1800)  # Run every 30 minutes

# Helper functions
def save_upload_file(upload_file: UploadFile) -> Path:
    """Save uploaded file to temporary directory"""
    file_id = str(uuid.uuid4())
    file_extension = Path(upload_file.filename).suffix
    file_path = UPLOAD_DIR / f"{file_id}{file_extension}"
    
    with open(file_path, "wb") as f:
        shutil.copyfileobj(upload_file.file, f)
    
    return file_path

def pdf_to_word(pdf_path: Path, output_path: Path):
    """Convert PDF to Word"""
    cv = Converter(str(pdf_path))
    cv.convert(str(output_path))
    cv.close()

def word_to_pdf(docx_path: Path, output_path: Path):
    """Convert Word to PDF using reportlab"""
    doc = Document(docx_path)
    
    pdf_canvas = canvas.Canvas(str(output_path), pagesize=letter)
    width, height = letter
    y_position = height - 50
    
    for para in doc.paragraphs:
        if para.text.strip():
            pdf_canvas.drawString(50, y_position, para.text[:100])
            y_position -= 20
            if y_position < 50:
                pdf_canvas.showPage()
                y_position = height - 50
    
    pdf_canvas.save()

def image_to_pdf(image_path: Path, output_path: Path):
    """Convert image to PDF"""
    image = Image.open(image_path)
    if image.mode == 'RGBA':
        image = image.convert('RGB')
    image.save(output_path, 'PDF', resolution=100.0)

def pdf_to_image(pdf_path: Path, output_path: Path, page_num: int = 0):
    """Convert first page of PDF to image using PyMuPDF"""
    import fitz
    doc = fitz.open(pdf_path)
    page = doc[page_num]
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
    pix.save(output_path)
    doc.close()

def merge_pdfs(pdf_paths: List[Path], output_path: Path):
    """Merge multiple PDFs"""
    merger = PdfMerger()
    for pdf_path in pdf_paths:
        merger.append(str(pdf_path))
    merger.write(str(output_path))
    merger.close()

def split_pdf(pdf_path: Path, output_dir: Path, pages: List[int]):
    """Split PDF into separate pages"""
    reader = PdfReader(str(pdf_path))
    output_files = []
    
    for page_num in pages:
        if page_num < len(reader.pages):
            writer = PdfWriter()
            writer.add_page(reader.pages[page_num])
            output_file = output_dir / f"page_{page_num + 1}.pdf"
            with open(output_file, "wb") as f:
                writer.write(f)
            output_files.append(output_file)
    
    return output_files

def compress_pdf(pdf_path: Path, output_path: Path):
    """Compress PDF"""
    reader = PdfReader(str(pdf_path))
    writer = PdfWriter()
    
    for page in reader.pages:
        page.compress_content_streams()
        writer.add_page(page)
    
    with open(output_path, "wb") as f:
        writer.write(f)

def rotate_pdf(pdf_path: Path, output_path: Path, rotation: int):
    """Rotate PDF pages"""
    reader = PdfReader(str(pdf_path))
    writer = PdfWriter()
    
    for page in reader.pages:
        page.rotate(rotation)
        writer.add_page(page)
    
    with open(output_path, "wb") as f:
        writer.write(f)

def image_to_text_ocr(image_path: Path) -> str:
    """Extract text from image using OCR"""
    image = cv2.imread(str(image_path))
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    text = pytesseract.image_to_string(gray)
    return text

def excel_to_pdf(excel_path: Path, output_path: Path):
    """Convert Excel to PDF (simplified)"""
    wb = load_workbook(excel_path)
    ws = wb.active
    
    pdf_canvas = canvas.Canvas(str(output_path), pagesize=letter)
    y_position = 750
    
    for row in ws.iter_rows(values_only=True):
        row_text = ' | '.join([str(cell) if cell else '' for cell in row])
        pdf_canvas.drawString(50, y_position, row_text[:100])
        y_position -= 15
        if y_position < 50:
            pdf_canvas.showPage()
            y_position = 750
    
    pdf_canvas.save()

def convert_image_format(input_path: Path, output_path: Path, output_format: str):
    """Convert between image formats"""
    image = Image.open(input_path)
    if output_format.upper() == 'JPG' and image.mode == 'RGBA':
        image = image.convert('RGB')
    image.save(output_path, format=output_format.upper())

# API Routes
@api_router.get("/")
async def root():
    return {"message": "ConvertDocs API", "version": "1.0"}

# Blog endpoints
@api_router.post("/blog", response_model=BlogPost)
async def create_blog_post(post: BlogPostCreate):
    blog_dict = post.model_dump()
    blog_obj = BlogPost(**blog_dict)
    
    doc = blog_obj.model_dump()
    doc['published_at'] = doc['published_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.blog_posts.insert_one(doc)
    return blog_obj

@api_router.get("/blog", response_model=List[BlogPost])
async def get_blog_posts(category: Optional[str] = None):
    query = {"category": category} if category else {}
    posts = await db.blog_posts.find(query, {"_id": 0}).sort("published_at", -1).to_list(100)
    
    for post in posts:
        if isinstance(post['published_at'], str):
            post['published_at'] = datetime.fromisoformat(post['published_at'])
        if isinstance(post['updated_at'], str):
            post['updated_at'] = datetime.fromisoformat(post['updated_at'])
    
    return posts

@api_router.get("/blog/{slug}", response_model=BlogPost)
async def get_blog_post(slug: str):
    post = await db.blog_posts.find_one({"slug": slug}, {"_id": 0})
    if not post:
        raise HTTPException(status_code=404, detail="Blog post not found")
    
    if isinstance(post['published_at'], str):
        post['published_at'] = datetime.fromisoformat(post['published_at'])
    if isinstance(post['updated_at'], str):
        post['updated_at'] = datetime.fromisoformat(post['updated_at'])
    
    return post

# Contact endpoint
@api_router.post("/contact", response_model=ContactMessage)
async def create_contact_message(message: ContactMessageCreate):
    contact_dict = message.model_dump()
    contact_obj = ContactMessage(**contact_dict)
    
    doc = contact_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.contact_messages.insert_one(doc)
    return contact_obj

# Conversion endpoints
@api_router.post("/convert/pdf-to-word")
async def convert_pdf_to_word(file: UploadFile = File(...)):
    try:
        input_path = save_upload_file(file)
        output_path = input_path.with_suffix('.docx')
        
        pdf_to_word(input_path, output_path)
        
        response = FileResponse(
            path=output_path,
            media_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            filename=f"{Path(file.filename).stem}.docx"
        )
        
        # Schedule cleanup
        asyncio.create_task(cleanup_file(input_path, output_path))
        
        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

@api_router.post("/convert/word-to-pdf")
async def convert_word_to_pdf_endpoint(file: UploadFile = File(...)):
    try:
        input_path = save_upload_file(file)
        output_path = input_path.with_suffix('.pdf')
        
        word_to_pdf(input_path, output_path)
        
        response = FileResponse(
            path=output_path,
            media_type='application/pdf',
            filename=f"{Path(file.filename).stem}.pdf"
        )
        
        asyncio.create_task(cleanup_file(input_path, output_path))
        
        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

@api_router.post("/convert/image-to-pdf")
async def convert_image_to_pdf_endpoint(file: UploadFile = File(...)):
    try:
        input_path = save_upload_file(file)
        output_path = input_path.with_suffix('.pdf')
        
        image_to_pdf(input_path, output_path)
        
        response = FileResponse(
            path=output_path,
            media_type='application/pdf',
            filename=f"{Path(file.filename).stem}.pdf"
        )
        
        asyncio.create_task(cleanup_file(input_path, output_path))
        
        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

@api_router.post("/convert/pdf-to-image")
async def convert_pdf_to_image_endpoint(file: UploadFile = File(...)):
    try:
        input_path = save_upload_file(file)
        output_path = input_path.with_suffix('.jpg')
        
        pdf_to_image(input_path, output_path)
        
        response = FileResponse(
            path=output_path,
            media_type='image/jpeg',
            filename=f"{Path(file.filename).stem}.jpg"
        )
        
        asyncio.create_task(cleanup_file(input_path, output_path))
        
        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

@api_router.post("/convert/image-to-text")
async def convert_image_to_text_endpoint(file: UploadFile = File(...)):
    try:
        input_path = save_upload_file(file)
        
        text = image_to_text_ocr(input_path)
        
        asyncio.create_task(cleanup_file(input_path))
        
        return {"text": text, "success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR failed: {str(e)}")

@api_router.post("/convert/merge-pdf")
async def merge_pdf_endpoint(files: List[UploadFile] = File(...)):
    try:
        input_paths = [save_upload_file(file) for file in files]
        output_path = UPLOAD_DIR / f"{uuid.uuid4()}_merged.pdf"
        
        merge_pdfs(input_paths, output_path)
        
        response = FileResponse(
            path=output_path,
            media_type='application/pdf',
            filename="merged.pdf"
        )
        
        asyncio.create_task(cleanup_file(*input_paths, output_path))
        
        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Merge failed: {str(e)}")

@api_router.post("/convert/compress-pdf")
async def compress_pdf_endpoint(file: UploadFile = File(...)):
    try:
        input_path = save_upload_file(file)
        output_path = input_path.with_name(f"{input_path.stem}_compressed.pdf")
        
        compress_pdf(input_path, output_path)
        
        response = FileResponse(
            path=output_path,
            media_type='application/pdf',
            filename=f"{Path(file.filename).stem}_compressed.pdf"
        )
        
        asyncio.create_task(cleanup_file(input_path, output_path))
        
        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Compression failed: {str(e)}")

@api_router.post("/convert/rotate-pdf")
async def rotate_pdf_endpoint(file: UploadFile = File(...), rotation: int = Form(90)):
    try:
        input_path = save_upload_file(file)
        output_path = input_path.with_name(f"{input_path.stem}_rotated.pdf")
        
        rotate_pdf(input_path, output_path, rotation)
        
        response = FileResponse(
            path=output_path,
            media_type='application/pdf',
            filename=f"{Path(file.filename).stem}_rotated.pdf"
        )
        
        asyncio.create_task(cleanup_file(input_path, output_path))
        
        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Rotation failed: {str(e)}")

@api_router.post("/convert/excel-to-pdf")
async def convert_excel_to_pdf_endpoint(file: UploadFile = File(...)):
    try:
        input_path = save_upload_file(file)
        output_path = input_path.with_suffix('.pdf')
        
        excel_to_pdf(input_path, output_path)
        
        response = FileResponse(
            path=output_path,
            media_type='application/pdf',
            filename=f"{Path(file.filename).stem}.pdf"
        )
        
        asyncio.create_task(cleanup_file(input_path, output_path))
        
        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

@api_router.post("/convert/image-format")
async def convert_image_format_endpoint(
    file: UploadFile = File(...), 
    output_format: str = Form(...)
):
    try:
        input_path = save_upload_file(file)
        output_path = input_path.with_suffix(f".{output_format.lower()}")
        
        convert_image_format(input_path, output_path, output_format)
        
        media_types = {
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'webp': 'image/webp',
            'bmp': 'image/bmp'
        }
        
        response = FileResponse(
            path=output_path,
            media_type=media_types.get(output_format.lower(), 'image/jpeg'),
            filename=f"{Path(file.filename).stem}.{output_format.lower()}"
        )
        
        asyncio.create_task(cleanup_file(input_path, output_path))
        
        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

async def cleanup_file(*file_paths: Path):
    """Delete files after a delay"""
    await asyncio.sleep(60)
    for file_path in file_paths:
        try:
            if file_path.exists():
                file_path.unlink()
        except Exception as e:
            logging.error(f"Failed to delete {file_path}: {e}")

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    asyncio.create_task(cleanup_old_files())
    logger.info("ConvertDocs API started")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()